#!/usr/bin/env python3
"""
Dashboard de Entrega de Resultados
Procesa archivos XLS semanales y genera tablero de métricas en Excel + JSON.

Uso:
  python procesar.py                         # procesa todos los XLS en datos/
  python procesar.py --semana 24             # solo semana 24
  python procesar.py --semana 23 24          # semanas 23 y 24 combinadas
  python procesar.py --archivo datos/Semana_24.xls ...
"""

import sys
import os
import re
import json
import argparse
import datetime
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración ─────────────────────────────────────────────────────────────

SUCURSALES = {
    "LM": "López Mateos",
    "AC": "Acueducto",
    "CH": "Chamizal",
    "GP": "Guadalupe",
    "CM": "Clínica de la Mujer",
    "MP": "Medicina Preventiva",
    "LA": "Las Águilas",
}

ORDEN_UNIDADES = ["General", "LM", "AC", "CH", "GP", "CM", "MP", "LA"]

MODALIDADES = {
    "Cardiología":  {"depto": "Cardiología",               "lim_ag": 7, "lim_te": 7},
    "Densitometría":{"depto": "Densitometría",             "lim_ag": 2, "lim_te": 2},
    "Ecosonografía":{"depto": "Ecosonografía",             "lim_ag": 7, "lim_te": 2},
    "Mamografía":   {"depto": "Mamografía",                "lim_ag": 3, "lim_te": 3},
    "Radiología":   {"depto": "Radiología",                "lim_ag": 2, "lim_te": 2},
    "RM":           {"depto": "Resonancia Magnética",      "lim_ag": 5, "lim_te": 5},
    "Tomografía":   {"depto": "Tomografía Axial Computada","lim_ag": 3, "lim_te": 3},
}

DEPTO_A_MOD = {v["depto"]: k for k, v in MODALIDADES.items()}

RM_48H  = ["tórax a muslo", "bodyscan", "valoración funcional", "oncológica",
           "funcional", "tractografía", "difusión", "perfusión"]
RM_5H   = ["columna", "cervical", "dorsal", "lumbar", "hombro", "cadera",
           "sacroilíaca", "tobillo", "codo", "rodilla"]
RM_5H_EXCL = ["muñeca", "mano"]

DATOS_DIR  = Path(__file__).parent / "datos"
OUTPUT_DIR = Path(__file__).parent / "output"

# ── Colores tema ──────────────────────────────────────────────────────────────

COLOR_AZUL_OSCURO = "1F3864"
COLOR_AZUL_MEDIO  = "2E5FA3"
COLOR_AZUL_CLARO  = "BDD7EE"
COLOR_VERDE       = "70AD47"
COLOR_AMARILLO    = "FFD966"
COLOR_ROJO        = "FF0000"
COLOR_GRIS        = "D9D9D9"
COLOR_BLANCO      = "FFFFFF"
COLOR_ENCABEZADO  = "203864"
COLOR_FILA_PAR    = "EBF3FB"
COLOR_VERDE_CLARO = "E2EFDA"


# ── Utilidades ────────────────────────────────────────────────────────────────

def parse_af_hours(val):
    """Convierte string A-F ' HH:MM' o '-HH:MM' a float horas. None si vacío."""
    if not val or not isinstance(val, str):
        return None
    val = val.strip()
    if not val:
        return None
    neg = val.startswith("-")
    val = val.lstrip("-").strip()
    m = re.match(r"(\d+):(\d+)", val)
    if not m:
        return None
    hours = int(m.group(1)) + int(m.group(2)) / 60
    return -hours if neg else hours


def classify_rm(estudio: str) -> str:
    """Clasifica un estudio RM en '5h', '24h' o '48h'."""
    s = estudio.lower()
    if any(kw in s for kw in RM_48H):
        return "48h"
    if any(kw in s for kw in RM_5H) and not any(ex in s for ex in RM_5H_EXCL):
        return "5h"
    return "24h"


def semana_de_archivo(path: Path) -> int | None:
    """Extrae número de semana del nombre del archivo (Semana_NN.xls)."""
    m = re.search(r"[Ss]emana[_\s]*(\d+)", path.stem)
    return int(m.group(1)) if m else None


def fmt_horas(val):
    """Formatea horas float a 'HH:MM' para visualización."""
    if val is None:
        return "-"
    neg = val < 0
    val = abs(val)
    h = int(val)
    mn = int(round((val - h) * 60))
    return f"{'-' if neg else ''}{h:02d}:{mn:02d}"


# ── Interpretaciones (EXA / SIO) ─────────────────────────────────────────────

INTERP_DIR = Path(__file__).parent / "datos" / "interpretaciones"


def _semana_interp(path: Path) -> int | None:
    m = re.search(r"[Ss]emana[_\s]*(\d+)|[Ww]eek[_\s]*(\d+)", path.stem)
    if m:
        return int(m.group(1) or m.group(2))
    return None


def leer_exa(path: Path) -> dict:
    """Lee un archivo EXA .xlsx y retorna {usuario: total_estudios}."""
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    usuarios = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        user, _, _, _, count = (list(row) + [None]*5)[:5]
        if user:
            usuarios[str(user).strip()] = int(count) if count else 0
    return usuarios


def leer_sio(path: Path) -> dict:
    """Lee un archivo SIO .xls y retorna {usuario: total_estudios}."""
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    # Fila 3: nombres de usuario (cols 1 en adelante, hasta 'Total' o 'sio')
    usuarios_row = ws.row_values(3)
    # La fila de totales no está en un índice fijo (varía según las categorías
    # presentes en cada semana). Se localiza buscando "Total" en la columna 0.
    totales_row = None
    for i in range(ws.nrows):
        if str(ws.cell_value(i, 0)).strip().lower() == "total":
            totales_row = ws.row_values(i)
            break
    if totales_row is None:
        totales_row = ws.row_values(ws.nrows - 1)
    resultado = {}
    for i, usr in enumerate(usuarios_row):
        if not usr or usr in ("Total", "sio"):
            continue
        total = totales_row[i] if i < len(totales_row) else 0
        if total:
            resultado[str(usr).strip()] = int(total)
    return resultado


def leer_todas_interpretaciones() -> dict:
    """
    Lee todos los archivos EXA y SIO de datos/interpretaciones/.
    Retorna dict con claves 'exa' y 'sio', cada una con:
      { 'acumulado': {usuario: n}, 'por_semana': {semana: {usuario: n}} }
    """
    resultado = {
        "exa": {"acumulado": {}, "por_semana": {}},
        "sio": {"acumulado": {}, "por_semana": {}},
    }
    if not INTERP_DIR.exists():
        return resultado

    for path in sorted(INTERP_DIR.iterdir()):
        sem = _semana_interp(path)
        nombre = path.name.lower()
        if "exa" in nombre and path.suffix in (".xlsx", ".xls"):
            datos = leer_exa(path) if path.suffix == ".xlsx" else {}
            sistema = "exa"
        elif "sio" in nombre and path.suffix in (".xls", ".xlsx"):
            datos = leer_sio(path) if path.suffix == ".xls" else {}
            sistema = "sio"
        else:
            continue

        if sem:
            resultado[sistema]["por_semana"][str(sem)] = datos

        # Acumular
        for usr, n in datos.items():
            resultado[sistema]["acumulado"][usr] = (
                resultado[sistema]["acumulado"].get(usr, 0) + n
            )

    return resultado


# ── Chequeos de transcripción / entrega (.ods / .xlsx) ───────────────────────

CHEQUEOS_DIR = Path(__file__).parent / "datos" / "chequeos"

# Umbrales (días) para el total Realización → Entregado
CHK_LIM_VERDE    = 7
CHK_LIM_AMARILLO = 10

# Objetivos de cumplimiento solicitados
OBJ_ENTREGA_REAL   = 95   # % entregado en/antes de la Fecha entrega real
OBJ_REALIZ_TERM    = 98   # % con Realización → Terminado ≤ 4 días
LIM_REALIZ_TERM    = 4    # días límite para la transcripción

# Ventana de saneamiento: descarta diferencias de días imposibles
# (errores de captura con años mal escritos, p. ej. 0226 → diferencias de miles de días).
DIAS_MIN_VALIDO = -60
DIAS_MAX_VALIDO = 365

# Normalización de nombres (unifica variantes/typos a un nombre canónico)
INTERNISTAS_CANON = ["Oropeza", "Castellón", "Santoyo", "Toriz", "Blanca"]
TRANSCRIPTORES_CANON = ["Casandra", "Beto", "Mayra", "Graciela", "Humberto"]

_INTERNISTA_ALIAS = {
    "castellon":     "Castellón",
    "castellón":     "Castellón",
    "dra. blanca":   "Blanca",
    "dra blanca":    "Blanca",
    "blanca":        "Blanca",
    "oropeza":       "Oropeza",
    "santoyo":       "Santoyo",
    "toriz":         "Toriz",
}

_TRANSCRIPTOR_ALIAS = {
    "casandra":  "Casandra",
    "beto":      "Beto",
    "mayra":     "Mayra",
    "graciela":  "Graciela",
    "humberto":  "Humberto",
}


def _norm_internista(nombre: str) -> str:
    """Devuelve el internista canónico o 'Otros' para valores fuera de la lista."""
    if not nombre:
        return "(sin asignar)"
    clave = str(nombre).strip().lower()
    if not clave or clave in ("-", "?", "x"):
        return "(sin asignar)"
    return _INTERNISTA_ALIAS.get(clave, "Otros")


def _norm_transcriptor(nombre: str) -> str:
    """Devuelve el transcriptor canónico o 'Otros' para valores fuera de la lista."""
    if not nombre:
        return "(sin asignar)"
    clave = str(nombre).strip().lower()
    if not clave or clave in ("-", "sin transcriptor"):
        return "(sin asignar)"
    return _TRANSCRIPTOR_ALIAS.get(clave, "Otros")


def _dias_valido(d):
    """True si la diferencia de días cae en un rango plausible (descarta typos)."""
    return d is not None and DIAS_MIN_VALIDO <= d <= DIAS_MAX_VALIDO

_ODS_NS = {
    "table":  "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
}


def _ods_q(tag: str) -> str:
    prefijo, nombre = tag.split(":")
    return "{%s}%s" % (_ODS_NS[prefijo], nombre)


def _domingos_en_rango(a, b):
    """Número de domingos en el intervalo (a, b]  (a excluido, b incluido)."""
    inicio = a + datetime.timedelta(days=1)
    if inicio > b:
        return 0
    # Primer domingo ≥ inicio (Python: lunes=0 … domingo=6)
    primer = inicio + datetime.timedelta(days=(6 - inicio.weekday()) % 7)
    if primer > b:
        return 0
    return (b - primer).days // 7 + 1


def _dias(a, b):
    """
    Días hábiles transcurridos entre dos fechas (b − a), contando solo de
    lunes a sábado: los domingos NO se cuentan. None si falta alguna fecha.
    """
    if not (a and b):
        return None
    if b >= a:
        return (b - a).days - _domingos_en_rango(a, b)
    # Diferencia negativa (errores de captura): simétrico
    return -((a - b).days - _domingos_en_rango(b, a))


def _parse_fecha_chequeo(val):
    """Parsea una fecha de chequeos: ISO, dd/mm/aaaa y typos comunes (24/062026)."""
    if not val:
        return None
    s = str(val).strip()
    if not s or s == "-":
        return None
    # Formato ISO (lo que entrega LibreOffice en office:date-value): YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        s2 = s.replace("-", "/")
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s2)          # dd/mm/aaaa
        if not m:
            m = re.match(r"(\d{1,2})/(\d{1,2})(\d{4})$", s2)       # typo: dd/mmaaaa
        if not m:
            return None
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(y, mo, d)
    except ValueError:
        return None


def leer_chequeos_ods(path: Path) -> list[dict]:
    """Lee un .ods de chequeos y retorna registros con fechas y días por proceso."""
    z = zipfile.ZipFile(str(path))
    root = ET.fromstring(z.read("content.xml"))
    sheet = next(root.iter(_ods_q("table:table")), None)
    if sheet is None:
        return []

    filas = []
    for r in sheet.iter(_ods_q("table:table-row")):
        celdas = []
        for c in r.iter(_ods_q("table:table-cell")):
            rep = int(c.get(_ods_q("table:number-columns-repeated"), "1") or 1)
            date_v = c.get(_ods_q("office:date-value"))
            num_v  = c.get(_ods_q("office:value"))
            txt    = "".join(c.itertext())
            celdas.extend([date_v or num_v or txt] * rep)
        while celdas and celdas[-1] in ("", None):
            celdas.pop()
        filas.append(celdas)

    sem = _semana_interp(path) or semana_de_archivo(path)
    registros = []
    for fila in filas[1:]:  # primera fila = encabezados
        if not fila or not str(fila[0]).strip():
            continue

        def g(i):
            return str(fila[i]).strip() if i < len(fila) and fila[i] is not None else ""

        realizacion  = _parse_fecha_chequeo(g(0))
        semaforo     = _parse_fecha_chequeo(g(1))
        entrega_real = _parse_fecha_chequeo(g(2))
        terminado    = _parse_fecha_chequeo(g(3))
        entregado    = _parse_fecha_chequeo(g(4))

        registros.append({
            "semana":           sem,
            "realizacion":      realizacion,
            "semaforo":         semaforo,
            "entrega_real":     entrega_real,
            "terminado":        terminado,
            "entregado":        entregado,
            "entrega_digital":  g(5),
            "medico":           g(6),
            "estado":           g(7),
            "transcriptor":     g(8),
            "audio":            g(9),
            "notas":            g(10),
            # Días por proceso
            "d_realiz_semaforo":      _dias(realizacion, semaforo),
            "d_realiz_terminado":     _dias(realizacion, terminado),
            "d_terminado_entregado":  _dias(terminado, entregado),
            "d_realiz_entregado":     _dias(realizacion, entregado),  # ← TOTAL
            "d_entregado_vs_real":    _dias(entrega_real, entregado), # + = retraso vs compromiso
        })
    return registros


def leer_chequeos_xlsx(path: Path) -> list[dict]:
    """
    Lee un .xlsx de chequeos (formato combinado con columna 'Cita' inicial) y
    retorna registros con el mismo esquema que leer_chequeos_ods.

    La semana de cada registro se deriva de la Fecha de realización (semana ISO),
    porque un archivo combinado cubre varias semanas.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    filas = list(ws.iter_rows(min_row=1, values_only=True))
    if not filas:
        return []

    # Mapa de columnas por encabezado (tolerante a acentos/espacios).
    encab = [str(c).strip().lower() if c is not None else "" for c in filas[0]]

    def col(*claves):
        for i, h in enumerate(encab):
            if any(k in h for k in claves):
                return i
        return None

    c_realiz = col("realiz")
    c_semaf  = col("semáforo", "semaforo")
    c_ereal  = col("entrega real")
    c_term   = col("terminado")
    c_entr   = col("entregado")
    c_digital= col("entrega digital", "digital")
    c_medico = col("internista", "médico", "medico")
    c_estado = col("estado")
    c_transc = col("transcriptor")
    c_audio  = col("audio")

    def celda(fila, idx):
        if idx is None or idx >= len(fila):
            return None
        return fila[idx]

    def fecha(v):
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
        return _parse_fecha_chequeo(v)

    def texto(v):
        return str(v).strip() if v is not None else ""

    registros = []
    for fila in filas[1:]:
        if not fila or all(c is None or str(c).strip() == "" for c in fila):
            continue

        realizacion  = fecha(celda(fila, c_realiz))
        semaforo     = fecha(celda(fila, c_semaf))
        entrega_real = fecha(celda(fila, c_ereal))
        terminado    = fecha(celda(fila, c_term))
        entregado    = fecha(celda(fila, c_entr))

        # Semana ISO y mes a partir de la fecha de realización.
        sem = realizacion.isocalendar()[1] if realizacion else None
        mes = realizacion.strftime("%Y-%m") if realizacion else None

        registros.append({
            "semana":           sem,
            "mes":              mes,
            "realizacion":      realizacion,
            "semaforo":         semaforo,
            "entrega_real":     entrega_real,
            "terminado":        terminado,
            "entregado":        entregado,
            "entrega_digital":  texto(celda(fila, c_digital)),
            "medico":           texto(celda(fila, c_medico)),
            "estado":           texto(celda(fila, c_estado)),
            "transcriptor":     texto(celda(fila, c_transc)),
            "audio":            texto(celda(fila, c_audio)),
            "notas":            "",
            "d_realiz_semaforo":      _dias(realizacion, semaforo),
            "d_realiz_terminado":     _dias(realizacion, terminado),
            "d_terminado_entregado":  _dias(terminado, entregado),
            "d_realiz_entregado":     _dias(realizacion, entregado),
            "d_entregado_vs_real":    _dias(entrega_real, entregado),
        })
    return registros


def leer_todos_chequeos() -> list[dict]:
    """Lee todos los .ods y .xlsx de datos/chequeos/."""
    if not CHEQUEOS_DIR.exists():
        return []
    regs = []
    for path in sorted(CHEQUEOS_DIR.iterdir()):
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() == ".ods":
            regs.extend(leer_chequeos_ods(path))
        elif path.suffix.lower() == ".xlsx":
            regs.extend(leer_chequeos_xlsx(path))
    # Asegura que todo registro tenga 'mes' (los .ods no lo traen)
    for r in regs:
        if "mes" not in r or r["mes"] is None:
            r["mes"] = r["realizacion"].strftime("%Y-%m") if r.get("realizacion") else None
    return regs


# ── Lectura de XLS ────────────────────────────────────────────────────────────

def leer_xls(filepath: Path, semana: int | None = None) -> list[dict]:
    """Lee un XLS y retorna lista de dicts. semana proviene del nombre del archivo."""
    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_index(0)
    if semana is None:
        semana = semana_de_archivo(filepath)

    rows = []
    for i in range(9, ws.nrows):
        row = ws.row_values(i)
        estudio  = str(row[2]).strip() if row[2] else ""
        depto    = str(row[3]).strip() if row[3] else ""
        sucursal = str(row[44]).strip() if row[44] else ""

        if not depto or not sucursal:
            continue

        ag = row[21]
        ag = float(ag) if ag else None

        fecha_alta_raw   = row[19]
        fecha_recibe_raw = row[39]
        af_str = str(row[41]).strip() if row[41] else ""

        te = None
        if (isinstance(fecha_recibe_raw, float) and isinstance(fecha_alta_raw, float)
                and fecha_recibe_raw and fecha_alta_raw):
            te = (fecha_recibe_raw - fecha_alta_raw) * 24

        af_hours = parse_af_hours(af_str)

        fecha = None
        hora_alta = None
        if isinstance(fecha_alta_raw, float) and fecha_alta_raw:
            dt = xlrd.xldate_as_datetime(fecha_alta_raw, wb.datemode)
            fecha = dt.date()
            hora_alta = dt.hour

        # Tiempos de proceso adicionales (en horas si son float, else None)
        gc = float(row[28]) if row[28] and isinstance(row[28], (int, float)) else None
        gf = float(row[38]) if row[38] and isinstance(row[38], (int, float)) else None
        usuario_recibe = str(row[40]).strip() if row[40] else ""

        rows.append({
            "estudio":         estudio,
            "depto":           depto,
            "sucursal":        sucursal,
            "ag":              ag,
            "af":              af_hours,
            "te":              te,
            "gc":              gc,   # alta → transcripción
            "gf":              gf,   # alta → recibe entrega
            "usuario_recibe":  usuario_recibe,
            "fecha":           fecha,
            "hora_alta":       hora_alta,
            "semana":          semana,
            "archivo":         filepath.name,
        })
    return rows


# ── Cálculo de métricas ───────────────────────────────────────────────────────

def es_incumplimiento(af):
    """Incumplimiento real: -24 < af < 0."""
    return af is not None and -24 < af < 0


def calcular_metricas(rows: list[dict]) -> dict:
    """Calcula todas las métricas para un conjunto de filas."""
    modal_rows = [r for r in rows if r["depto"] in DEPTO_A_MOD]
    total_modal = len(modal_rows)
    incump = sum(1 for r in modal_rows if es_incumplimiento(r["af"]))
    indicador = ((total_modal - incump) / total_modal * 100) if total_modal > 0 else None

    result = {
        "indicador":          indicador,
        "estudios":           total_modal,
        "incumplimiento":     incump,
        "modalidades_total":  len(rows),
    }

    for mod, cfg in MODALIDADES.items():
        depto   = cfg["depto"]
        lim_ag  = cfg["lim_ag"]
        lim_te  = cfg["lim_te"]

        mod_rows = [r for r in rows if r["depto"] == depto]
        ok_rows  = [r for r in mod_rows if r["ag"] is not None and r["ag"] <= lim_ag]
        count    = len(ok_rows)

        te_vals  = [r["te"] for r in ok_rows
                    if r["te"] is not None and 0 <= r["te"] <= lim_te]
        te_avg   = (sum(te_vals) / len(te_vals)) if te_vals else None

        result[mod]          = count
        result[f"TE_{mod}"]  = te_avg

        if mod == "RM":
            result["RM_5h"]  = sum(1 for r in ok_rows if classify_rm(r["estudio"]) == "5h")
            result["RM_24h"] = sum(1 for r in ok_rows if classify_rm(r["estudio"]) == "24h")
            result["RM_48h"] = sum(1 for r in ok_rows if classify_rm(r["estudio"]) == "48h")

    return result


def generar_tabla(rows: list[dict], fecha_reporte=None) -> list[dict]:
    """Genera lista de filas del dashboard (una por unidad)."""
    tabla = []
    for unidad in ORDEN_UNIDADES:
        if unidad == "General":
            subset = rows
            fecha  = str(fecha_reporte) if fecha_reporte else ""
        else:
            nombre_suc = SUCURSALES[unidad]
            subset = [r for r in rows if r["sucursal"] == nombre_suc]
            fecha  = ""

        m = calcular_metricas(subset)
        tabla.append({"Unidad": unidad, "Fecha": fecha, **m})
    return tabla


# ── Escritura Excel ───────────────────────────────────────────────────────────

def thin_border():
    t = Side(style="thin", color="AAAAAA")
    return Border(left=t, right=t, top=t, bottom=t)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def bold_font(size=11, color=COLOR_BLANCO):
    return Font(bold=True, size=size, color=color)


def normal_font(size=10, color="000000"):
    return Font(size=size, color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center")


COLUMNAS_DASH = [
    ("Fecha",             16),
    ("Unidad",            10),
    ("Indicador %",       12),
    ("Estudios",          10),
    ("Incumplimiento",    14),
    ("Total Registros",   13),
    ("Cardiología",       12),
    ("TE Cardiología",    14),
    ("Densitometría",     13),
    ("TE Densitometría",  15),
    ("Ecosonografía",     13),
    ("TE Ecosonografía",  15),
    ("Mamografía",        12),
    ("TE Mamografía",     14),
    ("Radiología",        12),
    ("TE Radiología",     14),
    ("RM",                8),
    ("TE RM",             10),
    ("RM 5h",             9),
    ("RM 24h",            9),
    ("RM 48h",            9),
    ("Tomografía",        12),
    ("TE Tomografía",     14),
]

# mapa de nombre de columna → clave en el dict de métricas
COL_KEY = {
    "Indicador %":      "indicador",
    "Estudios":         "estudios",
    "Incumplimiento":   "incumplimiento",
    "Total Registros":  "modalidades_total",
    "Cardiología":      "Cardiología",
    "TE Cardiología":   "TE_Cardiología",
    "Densitometría":    "Densitometría",
    "TE Densitometría": "TE_Densitometría",
    "Ecosonografía":    "Ecosonografía",
    "TE Ecosonografía": "TE_Ecosonografía",
    "Mamografía":       "Mamografía",
    "TE Mamografía":    "TE_Mamografía",
    "Radiología":       "Radiología",
    "TE Radiología":    "TE_Radiología",
    "RM":               "RM",
    "TE RM":            "TE_RM",
    "RM 5h":            "RM_5h",
    "RM 24h":           "RM_24h",
    "RM 48h":           "RM_48h",
    "Tomografía":       "Tomografía",
    "TE Tomografía":    "TE_Tomografía",
}


def escribir_hoja_dashboard(ws_dash, titulo: str, tabla: list[dict]):
    """Escribe la hoja Dashboard con formato completo."""
    # — Título —
    ws_dash.merge_cells(f"A1:{get_column_letter(len(COLUMNAS_DASH))}1")
    celda = ws_dash.cell(1, 1, titulo)
    celda.fill      = fill(COLOR_AZUL_OSCURO)
    celda.font      = Font(bold=True, size=14, color=COLOR_BLANCO)
    celda.alignment = center()
    ws_dash.row_dimensions[1].height = 30

    # — Encabezados —
    for col_idx, (nombre, ancho) in enumerate(COLUMNAS_DASH, 1):
        celda = ws_dash.cell(2, col_idx, nombre)
        celda.fill      = fill(COLOR_AZUL_MEDIO)
        celda.font      = bold_font(10)
        celda.alignment = center()
        celda.border    = thin_border()
        ws_dash.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws_dash.row_dimensions[2].height = 40

    # — Filas de datos —
    for fila_idx, fila in enumerate(tabla, 3):
        es_general = fila["Unidad"] == "General"
        color_fondo = COLOR_AZUL_CLARO if es_general else (
            COLOR_FILA_PAR if fila_idx % 2 == 0 else COLOR_BLANCO
        )

        for col_idx, (nombre, _) in enumerate(COLUMNAS_DASH, 1):
            if nombre == "Fecha":
                val = fila.get("Fecha", "")
            elif nombre == "Unidad":
                val = fila["Unidad"]
            else:
                key  = COL_KEY.get(nombre)
                val  = fila.get(key)

            # Formateo de valores
            if nombre == "Indicador %":
                texto = f"{val:.1f}%" if val is not None else "-"
                # Color según cumplimiento
                if val is not None:
                    color_ind = COLOR_VERDE if val >= 95 else (
                        COLOR_AMARILLO if val >= 90 else "FF4444")
                else:
                    color_ind = None
            elif nombre.startswith("TE "):
                texto = fmt_horas(val)
                color_ind = None
            elif val is None:
                texto = "-"
                color_ind = None
            elif isinstance(val, float) and not nombre.startswith("RM"):
                texto = f"{val:.0f}"
                color_ind = None
            else:
                texto = str(val) if val is not None else "-"
                color_ind = None

            celda = ws_dash.cell(fila_idx, col_idx, texto)
            celda.fill      = fill(color_fondo)
            celda.border    = thin_border()
            celda.alignment = center() if col_idx > 2 else left()

            if nombre == "Indicador %" and color_ind:
                celda.font = Font(bold=True, size=11, color=color_ind)
            elif es_general:
                celda.font = Font(bold=True, size=10, color="1F3864")
            else:
                celda.font = normal_font()

        ws_dash.row_dimensions[fila_idx].height = 22

    # Freeze encabezado
    ws_dash.freeze_panes = "C3"


def escribir_hoja_detalle(ws_det, rows: list[dict]):
    """Hoja con detalle de registros usados en el cálculo."""
    encabezados = ["Semana", "Fecha", "Sucursal", "Departamento", "Estudio",
                   "A-G (h)", "A-F", "TE (h)", "Incumplimiento", "Clasificación RM"]
    anchos      = [9, 12, 18, 22, 50, 10, 10, 10, 14, 16]

    for col_idx, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
        celda = ws_det.cell(1, col_idx, enc)
        celda.fill      = fill(COLOR_ENCABEZADO)
        celda.font      = bold_font(10)
        celda.alignment = center()
        celda.border    = thin_border()
        ws_det.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws_det.row_dimensions[1].height = 30

    for fila_idx, r in enumerate(rows, 2):
        incump_val = "Sí" if es_incumplimiento(r["af"]) else "No"
        rm_cls     = classify_rm(r["estudio"]) if r["depto"] == "Resonancia Magnética" else ""
        color_fondo = COLOR_FILA_PAR if fila_idx % 2 == 0 else COLOR_BLANCO

        valores = [
            r.get("semana", ""),
            str(r["fecha"]) if r["fecha"] else "",
            r["sucursal"],
            r["depto"],
            r["estudio"],
            f"{r['ag']:.1f}" if r["ag"] is not None else "-",
            fmt_horas(r["af"]),
            f"{r['te']:.2f}" if r["te"] is not None else "-",
            incump_val,
            rm_cls,
        ]
        for col_idx, val in enumerate(valores, 1):
            celda = ws_det.cell(fila_idx, col_idx, val)
            celda.fill      = fill(color_fondo)
            celda.font      = normal_font(9)
            celda.alignment = left()
            celda.border    = thin_border()
            if incump_val == "Sí" and col_idx == 9:
                celda.font = Font(size=9, color=COLOR_ROJO, bold=True)

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions


def escribir_hoja_semanas(ws_sem, todas_las_filas: list[dict], semanas: list[int]):
    """Resumen por semana: una sub-tabla por semana."""
    fila_actual = 1
    for semana in sorted(semanas):
        rows_sem = [r for r in todas_las_filas if r.get("semana") == semana]
        if not rows_sem:
            continue
        fechas = [r["fecha"] for r in rows_sem if r["fecha"]]
        rango  = f"{min(fechas)} al {max(fechas)}" if fechas else ""
        titulo = f"Semana {semana}  —  {rango}"

        tabla = generar_tabla(rows_sem, max(fechas) if fechas else None)

        # Escribir subtítulo
        ws_sem.merge_cells(
            start_row=fila_actual, start_column=1,
            end_row=fila_actual, end_column=len(COLUMNAS_DASH))
        celda = ws_sem.cell(fila_actual, 1, titulo)
        celda.fill      = fill(COLOR_AZUL_MEDIO)
        celda.font      = bold_font(11)
        celda.alignment = center()
        ws_sem.row_dimensions[fila_actual].height = 25
        fila_actual += 1

        # Encabezados
        for col_idx, (nombre, ancho) in enumerate(COLUMNAS_DASH, 1):
            celda = ws_sem.cell(fila_actual, col_idx, nombre)
            celda.fill      = fill(COLOR_AZUL_CLARO)
            celda.font      = Font(bold=True, size=9, color="1F3864")
            celda.alignment = center()
            celda.border    = thin_border()
            ws_sem.column_dimensions[get_column_letter(col_idx)].width = ancho
        ws_sem.row_dimensions[fila_actual].height = 30
        fila_actual += 1

        # Datos
        for fila in tabla:
            es_general  = fila["Unidad"] == "General"
            color_fondo = COLOR_VERDE_CLARO if es_general else (
                COLOR_FILA_PAR if fila_actual % 2 == 0 else COLOR_BLANCO)

            for col_idx, (nombre, _) in enumerate(COLUMNAS_DASH, 1):
                if nombre == "Fecha":
                    val = fila.get("Fecha", "")
                elif nombre == "Unidad":
                    val = fila["Unidad"]
                else:
                    key = COL_KEY.get(nombre)
                    val = fila.get(key)

                if nombre == "Indicador %":
                    texto = f"{val:.1f}%" if val is not None else "-"
                elif nombre.startswith("TE "):
                    texto = fmt_horas(val)
                elif val is None:
                    texto = "-"
                else:
                    texto = str(val) if not isinstance(val, float) else f"{val:.0f}"

                celda = ws_sem.cell(fila_actual, col_idx, texto)
                celda.fill      = fill(color_fondo)
                celda.border    = thin_border()
                celda.alignment = center() if col_idx > 2 else left()
                celda.font      = (Font(bold=True, size=9, color="1F3864")
                                   if es_general else Font(size=9))

            ws_sem.row_dimensions[fila_actual].height = 18
            fila_actual += 1

        fila_actual += 2  # espacio entre semanas


def escribir_hoja_chequeos(ws, registros: list[dict]):
    """Hoja con los días que pasa cada estudio en cada proceso (transcripción → entrega)."""
    PROCESOS = [
        ("Realización → Semáforo (objetivo)",            "d_realiz_semaforo"),
        ("Realización → Terminado (transcripción)",      "d_realiz_terminado"),
        ("Terminado → Entregado (entrega)",              "d_terminado_entregado"),
        ("Realización → Entregado (TOTAL)",              "d_realiz_entregado"),
        ("Entregado vs compromiso (+ = retraso)",        "d_entregado_vs_real"),
    ]

    columnas = [
        ("Semana", 8), ("Realización", 13), ("Semáforo", 13), ("Terminado", 13),
        ("Entregado", 13), ("Médico", 12), ("Transcriptor", 13), ("Estado", 12),
        ("Entrega Digital", 13),
        ("R→Terminado", 12), ("Term→Entreg.", 13),
        ("R→Entregado (TOTAL)", 18), ("R→Semáforo", 12), ("Δ vs compromiso", 14),
    ]
    n_cols = len(columnas)

    # — Título —
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws.cell(1, 1, "Días por Proceso — Transcripción y Entrega de Estudios")
    c.fill = fill(COLOR_AZUL_OSCURO)
    c.font = Font(bold=True, size=14, color=COLOR_BLANCO)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    fila = 3

    # — Resumen: días promedio por proceso —
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    c = ws.cell(fila, 1, "Resumen — días por proceso")
    c.fill = fill(COLOR_AZUL_MEDIO); c.font = bold_font(11); c.alignment = center()
    fila += 1

    for j, h in enumerate(["Proceso", "Promedio (días)", "Mín", "Máx", "N"], 1):
        cc = ws.cell(fila, j, h)
        cc.fill = fill(COLOR_AZUL_CLARO); cc.font = Font(bold=True, size=10, color="1F3864")
        cc.alignment = center(); cc.border = thin_border()
    fila += 1

    for etiqueta, key in PROCESOS:
        vals = [r[key] for r in registros if _dias_valido(r[key])]
        avg  = sum(vals) / len(vals) if vals else None
        destacado = key == "d_realiz_entregado"
        valores = [
            etiqueta,
            f"{avg:.1f}" if avg is not None else "-",
            str(min(vals)) if vals else "-",
            str(max(vals)) if vals else "-",
            len(vals),
        ]
        for j, v in enumerate(valores, 1):
            cc = ws.cell(fila, j, v)
            cc.border = thin_border()
            cc.alignment = left() if j == 1 else center()
            if destacado:
                cc.fill = fill(COLOR_AMARILLO); cc.font = Font(bold=True, size=10)
            else:
                cc.fill = fill(COLOR_BLANCO); cc.font = normal_font(10)
        fila += 1

    fila += 1

    # — Resumen por transcriptor (días totales R→Entregado) —
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    c = ws.cell(fila, 1, "Resumen por transcriptor — Realización → Entregado")
    c.fill = fill(COLOR_AZUL_MEDIO); c.font = bold_font(11); c.alignment = center()
    fila += 1

    for j, h in enumerate(["Transcriptor", "Entregados", "Prom. días", "Mín", "Máx"], 1):
        cc = ws.cell(fila, j, h)
        cc.fill = fill(COLOR_AZUL_CLARO); cc.font = Font(bold=True, size=10, color="1F3864")
        cc.alignment = center(); cc.border = thin_border()
    fila += 1

    por_tr = {}
    for r in registros:
        tr = r["transcriptor"] or "(sin asignar)"
        if tr in ("-",):
            tr = "(sin asignar)"
        por_tr.setdefault(tr, []).append(r["d_realiz_entregado"])

    for tr in sorted(por_tr):
        vals = [v for v in por_tr[tr] if v is not None]
        avg  = sum(vals) / len(vals) if vals else None
        valores = [
            tr,
            len(vals),
            f"{avg:.1f}" if avg is not None else "-",
            str(min(vals)) if vals else "-",
            str(max(vals)) if vals else "-",
        ]
        for j, v in enumerate(valores, 1):
            cc = ws.cell(fila, j, v)
            cc.border = thin_border()
            cc.alignment = left() if j == 1 else center()
            cc.fill = fill(COLOR_BLANCO); cc.font = normal_font(10)
        fila += 1

    fila += 1

    # — Encabezados del detalle —
    encab_fila = fila
    for col_idx, (nombre, ancho) in enumerate(columnas, 1):
        cc = ws.cell(encab_fila, col_idx, nombre)
        cc.fill = fill(COLOR_ENCABEZADO); cc.font = bold_font(9)
        cc.alignment = center(); cc.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[encab_fila].height = 30
    fila += 1

    def fdate(d):
        return d.isoformat() if d else "-"

    def fnum(n):
        return n if n is not None else "-"

    for r in registros:
        total = r["d_realiz_entregado"]
        valores = [
            r.get("semana") or "",
            fdate(r["realizacion"]),
            fdate(r["semaforo"]),
            fdate(r["terminado"]),
            fdate(r["entregado"]) if r["entregado"] else "En proceso",
            r["medico"],
            r["transcriptor"],
            r["estado"],
            r["entrega_digital"],
            fnum(r["d_realiz_terminado"]),
            fnum(r["d_terminado_entregado"]),
            fnum(total),
            fnum(r["d_realiz_semaforo"]),
            fnum(r["d_entregado_vs_real"]),
        ]
        color_fondo = COLOR_FILA_PAR if fila % 2 == 0 else COLOR_BLANCO
        for col_idx, v in enumerate(valores, 1):
            cc = ws.cell(fila, col_idx, v)
            cc.border = thin_border()
            cc.font = normal_font(9)
            cc.alignment = left() if col_idx <= 9 else center()
            cc.fill = fill(color_fondo)
            # Semáforo de color en la columna TOTAL (col 12)
            if col_idx == 12 and total is not None:
                if total <= CHK_LIM_VERDE:
                    cc.fill = fill(COLOR_VERDE_CLARO); cc.font = Font(size=9, bold=True, color="1F6B2E")
                elif total <= CHK_LIM_AMARILLO:
                    cc.fill = fill(COLOR_AMARILLO); cc.font = Font(size=9, bold=True, color="7A5C00")
                else:
                    cc.fill = fill("FFC7C7"); cc.font = Font(size=9, bold=True, color=COLOR_ROJO)
        ws.row_dimensions[fila].height = 16
        fila += 1

    ws.freeze_panes = ws.cell(encab_fila + 1, 1)
    ws.auto_filter.ref = (
        f"A{encab_fila}:{get_column_letter(n_cols)}{fila - 1}"
    )


# ── Exportar JSON ─────────────────────────────────────────────────────────────

def exportar_json(tabla_acum: list[dict], tabla_por_semana: dict,
                  rows: list[dict], ruta: Path):
    """Genera JSON de respaldo con todos los datos."""
    def clean(obj):
        if isinstance(obj, dict):
            return {k: clean(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [clean(v) for v in obj]
        if isinstance(obj, datetime.date):
            return obj.isoformat()
        if isinstance(obj, float) and obj != obj:  # NaN
            return None
        return obj

    data = {
        "generado":      datetime.datetime.now().isoformat(),
        "acumulado":     tabla_acum,
        "por_semana":    tabla_por_semana,
        "total_registros": len(rows),
    }
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(clean(data), f, ensure_ascii=False, indent=2, default=str)


def _resumen_periodo(registros: list[dict]) -> dict:
    """
    Calcula el paquete de métricas solicitadas para un subconjunto de registros
    (sirve para el acumulado y para cada semana / mes):

      • Entregado − Entrega real  → días de anticipación y % a tiempo (obj. 95%)
      • Realización → Terminado   → días de transcripción y % ≤ 4 días (obj. 98%)
      • Cantidad por internista
      • Cantidad por transcriptor
    """
    def stats(key):
        vals = [r[key] for r in registros if _dias_valido(r[key])]
        return {
            "avg": round(sum(vals) / len(vals), 1) if vals else None,
            "min": min(vals) if vals else None,
            "max": max(vals) if vals else None,
            "n":   len(vals),
        }

    # ── Entregado vs Fecha entrega real ──────────────────────────────────────
    # dias_antes = entrega_real − entregado  (positivo = entregado antes de la fecha)
    antes_vals = [-r["d_entregado_vs_real"] for r in registros
                  if _dias_valido(r["d_entregado_vs_real"])]
    a_tiempo = sum(1 for d in antes_vals if d >= 0)
    entrega_real = {
        "n":            len(antes_vals),
        "a_tiempo":     a_tiempo,
        "tarde":        len(antes_vals) - a_tiempo,
        "pct_a_tiempo": round(a_tiempo / len(antes_vals) * 100, 1) if antes_vals else None,
        "avg_antes":    round(sum(antes_vals) / len(antes_vals), 1) if antes_vals else None,
        "objetivo":     OBJ_ENTREGA_REAL,
    }
    dist_antes_map = {}
    for d in antes_vals:
        dist_antes_map[d] = dist_antes_map.get(d, 0) + 1
    entrega_real["dist"] = [{"dias": k, "n": dist_antes_map[k]}
                            for k in sorted(dist_antes_map)]

    # ── Realización → Terminado (transcripción) ──────────────────────────────
    rt_vals = [r["d_realiz_terminado"] for r in registros
               if _dias_valido(r["d_realiz_terminado"])]
    rt_ok = sum(1 for d in rt_vals if d <= LIM_REALIZ_TERM)
    realiz_term = {
        "n":          len(rt_vals),
        "ok":         rt_ok,
        "fuera":      len(rt_vals) - rt_ok,
        "pct_ok":     round(rt_ok / len(rt_vals) * 100, 1) if rt_vals else None,
        "avg":        round(sum(rt_vals) / len(rt_vals), 1) if rt_vals else None,
        "limite":     LIM_REALIZ_TERM,
        "objetivo":   OBJ_REALIZ_TERM,
    }
    dist_rt_map = {}
    for d in rt_vals:
        dist_rt_map[d] = dist_rt_map.get(d, 0) + 1
    realiz_term["dist"] = [{"dias": k, "n": dist_rt_map[k]}
                           for k in sorted(dist_rt_map)]

    # ── Cantidad por internista ──────────────────────────────────────────────
    int_cnt = {}
    for r in registros:
        int_cnt[_norm_internista(r["medico"])] = \
            int_cnt.get(_norm_internista(r["medico"]), 0) + 1
    orden_int = INTERNISTAS_CANON + ["Otros", "(sin asignar)"]
    internistas = [{"nombre": n, "n": int_cnt[n]}
                   for n in orden_int if int_cnt.get(n)]

    # ── Cantidad por transcriptor ────────────────────────────────────────────
    tr_cnt = {}
    for r in registros:
        tr_cnt[_norm_transcriptor(r["transcriptor"])] = \
            tr_cnt.get(_norm_transcriptor(r["transcriptor"]), 0) + 1
    orden_tr = TRANSCRIPTORES_CANON + ["Otros", "(sin asignar)"]
    transcriptores_cnt = [{"nombre": n, "n": tr_cnt[n]}
                          for n in orden_tr if tr_cnt.get(n)]

    return {
        "total":          len(registros),
        "entregados":     sum(1 for r in registros if r["entregado"]),
        "realiz_terminado":    stats("d_realiz_terminado"),
        "terminado_entregado": stats("d_terminado_entregado"),
        "realiz_entregado":    stats("d_realiz_entregado"),
        "entrega_real":   entrega_real,
        "realiz_term_obj": realiz_term,
        "internistas":    internistas,
        "transcriptores_cnt": transcriptores_cnt,
    }


_DOW_NOMBRES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]


def _patrones_chequeos(registros, por_semana, por_mes, resumen) -> dict:
    """
    Detecta patrones accionables: tendencia temporal, días pico de la semana,
    desempeño por transcriptor / internista y hallazgos automáticos para la
    toma de decisiones.
    """
    # ── Tendencia por semana y por mes (serie temporal) ──────────────────────
    def serie(dic, claves):
        out = []
        for k in claves:
            b = dic[k]
            out.append({
                "periodo":     k,
                "total":       b["total"],
                "pct_transc":  b["realiz_term_obj"]["pct_ok"],
                "avg_transc":  b["realiz_term_obj"]["avg"],
                "pct_entrega": b["entrega_real"]["pct_a_tiempo"],
            })
        return out

    tendencia_semana = serie(por_semana, sorted(por_semana, key=lambda x: int(x)))
    tendencia_mes    = serie(por_mes, sorted(por_mes))

    # ── Volumen por día de la semana (lun–sáb; domingo no se cuenta) ──────────
    dow = {i: 0 for i in range(6)}
    for r in registros:
        f = r.get("realizacion")
        if f and f.weekday() < 6:
            dow[f.weekday()] += 1
    por_dia = [{"dia": _DOW_NOMBRES[i], "n": dow[i]} for i in range(6)]

    # ── Desempeño por transcriptor / internista ──────────────────────────────
    def desempeno(group_fn, dkey, lim):
        g = {}
        for r in registros:
            v = r[dkey]
            if not _dias_valido(v):
                continue
            k = group_fn(r)
            g.setdefault(k, {"n": 0, "ok": 0, "suma": 0})
            g[k]["n"]   += 1
            g[k]["suma"] += v
            if v <= lim:
                g[k]["ok"] += 1
        filas = []
        for k, v in g.items():
            filas.append({
                "nombre": k, "n": v["n"],
                "pct_ok": round(v["ok"] / v["n"] * 100, 1) if v["n"] else None,
                "avg":    round(v["suma"] / v["n"], 1) if v["n"] else None,
            })
        return sorted(filas, key=lambda x: -x["n"])

    transc_perf = desempeno(lambda r: _norm_transcriptor(r["transcriptor"]),
                            "d_realiz_terminado", LIM_REALIZ_TERM)
    intern_perf = desempeno(lambda r: _norm_internista(r["medico"]),
                            "d_realiz_entregado", CHK_LIM_VERDE)

    # ── Hallazgos automáticos ────────────────────────────────────────────────
    hallazgos = []
    er = resumen["entrega_real"]
    rt = resumen["realiz_term_obj"]

    # Cuello de botella
    if rt["pct_ok"] is not None and er["pct_a_tiempo"] is not None:
        if rt["pct_ok"] < OBJ_REALIZ_TERM and er["pct_a_tiempo"] >= OBJ_ENTREGA_REAL:
            hallazgos.append({
                "tipo": "bad",
                "txt": f"El cuello de botella es la TRANSCRIPCIÓN: solo {rt['pct_ok']}% se "
                       f"termina en ≤{LIM_REALIZ_TERM} días (objetivo {OBJ_REALIZ_TERM}%), "
                       f"mientras que la entrega final cumple {er['pct_a_tiempo']}% "
                       f"(objetivo {OBJ_ENTREGA_REAL}%). El retraso se origina antes de la entrega.",
            })
        elif rt["pct_ok"] >= OBJ_REALIZ_TERM:
            hallazgos.append({
                "tipo": "good",
                "txt": f"Ambos objetivos se cumplen: transcripción {rt['pct_ok']}% "
                       f"(≥{OBJ_REALIZ_TERM}%) y entrega {er['pct_a_tiempo']}% (≥{OBJ_ENTREGA_REAL}%).",
            })

    # Semanas críticas (más de 8 registros y cumplimiento bajo)
    criticas = [s for s in tendencia_semana
                if s["total"] >= 8 and s["pct_transc"] is not None
                and s["pct_transc"] < 60]
    if criticas:
        criticas.sort(key=lambda x: x["pct_transc"])
        etqs = ", ".join(f"Sem {c['periodo']} ({c['pct_transc']}%)" for c in criticas[:4])
        hallazgos.append({
            "tipo": "bad",
            "txt": f"Semanas críticas en transcripción (<60% en ≤{LIM_REALIZ_TERM} días): {etqs}. "
                   f"Conviene revisar carga de trabajo o ausencias en esas semanas.",
        })

    # Tendencia entre primer y último mes
    if len(tendencia_mes) >= 2:
        m0, m1 = tendencia_mes[0], tendencia_mes[-1]
        if m0["pct_transc"] is not None and m1["pct_transc"] is not None:
            delta = round(m1["pct_transc"] - m0["pct_transc"], 1)
            if abs(delta) >= 5:
                hallazgos.append({
                    "tipo": "good" if delta > 0 else "warn",
                    "txt": f"La transcripción {'mejoró' if delta>0 else 'empeoró'} "
                           f"{abs(delta)} puntos de {m0['periodo']} ({m0['pct_transc']}%) "
                           f"a {m1['periodo']} ({m1['pct_transc']}%).",
                })

    # Día pico de realización
    if por_dia:
        pico = max(por_dia, key=lambda x: x["n"])
        total_dias = sum(x["n"] for x in por_dia) or 1
        hallazgos.append({
            "tipo": "info",
            "txt": f"El {pico['dia']} concentra el mayor volumen de estudios "
                   f"({pico['n']}, {round(pico['n']/total_dias*100)}% del total). "
                   f"La carga se acumula a fin de semana (jueves a sábado), lo que presiona "
                   f"la transcripción de los días siguientes.",
        })

    # Comparación entre transcriptores principales
    principales = [t for t in transc_perf
                   if t["nombre"] in TRANSCRIPTORES_CANON and t["n"] >= 20]
    if len(principales) >= 2:
        principales.sort(key=lambda x: -(x["pct_ok"] or 0))
        mejor, peor = principales[0], principales[-1]
        if (mejor["pct_ok"] or 0) - (peor["pct_ok"] or 0) >= 5:
            hallazgos.append({
                "tipo": "info",
                "txt": f"Entre transcriptores, {mejor['nombre']} cumple {mejor['pct_ok']}% "
                       f"en ≤{LIM_REALIZ_TERM} días vs {peor['nombre']} {peor['pct_ok']}% "
                       f"(volúmenes {mejor['n']} y {peor['n']}).",
            })

    return {
        "tendencia_semana": tendencia_semana,
        "tendencia_mes":    tendencia_mes,
        "por_dia":          por_dia,
        "transc_perf":      transc_perf,
        "intern_perf":      intern_perf,
        "hallazgos":        hallazgos,
    }


def _analisis_chequeos(registros: list[dict]) -> dict:
    """Resume los chequeos (días por proceso) para el dashboard web."""
    def stats(key):
        vals = [r[key] for r in registros if _dias_valido(r[key])]
        return {
            "avg": round(sum(vals) / len(vals), 1) if vals else None,
            "min": min(vals) if vals else None,
            "max": max(vals) if vals else None,
            "n":   len(vals),
        }

    procesos = {
        "realiz_semaforo":     stats("d_realiz_semaforo"),
        "realiz_terminado":    stats("d_realiz_terminado"),
        "terminado_entregado": stats("d_terminado_entregado"),
        "realiz_entregado":    stats("d_realiz_entregado"),
        "entregado_vs_real":   stats("d_entregado_vs_real"),
    }

    # Por transcriptor (sobre el total Realización → Entregado)
    por_tr = {}
    for r in registros:
        tr = r["transcriptor"] or "(sin asignar)"
        if tr == "-":
            tr = "(sin asignar)"
        por_tr.setdefault(tr, []).append(r["d_realiz_entregado"])
    transcriptores = []
    for tr in sorted(por_tr):
        vals = [v for v in por_tr[tr] if _dias_valido(v)]
        transcriptores.append({
            "transcriptor": tr,
            "entregados":   len(vals),
            "avg": round(sum(vals) / len(vals), 1) if vals else None,
            "min": min(vals) if vals else None,
            "max": max(vals) if vals else None,
        })
    transcriptores.sort(key=lambda x: -x["entregados"])

    # Distribución del total de días Realización → Entregado
    dist = {}
    for r in registros:
        d = r["d_realiz_entregado"]
        if _dias_valido(d):
            dist[d] = dist.get(d, 0) + 1
    distribucion = [{"dias": k, "n": dist[k]} for k in sorted(dist)]

    entregados = sum(1 for r in registros if r["entregado"])

    # ── Desglose por semana y por mes (métricas solicitadas) ─────────────────
    semanas_ord = sorted({r["semana"] for r in registros if r["semana"] is not None})
    meses_ord   = sorted({r.get("mes") for r in registros if r.get("mes")})

    por_semana = {str(s): _resumen_periodo([r for r in registros if r["semana"] == s])
                  for s in semanas_ord}
    por_mes    = {m: _resumen_periodo([r for r in registros if r.get("mes") == m])
                  for m in meses_ord}
    resumen_global = _resumen_periodo(registros)
    patrones = _patrones_chequeos(registros, por_semana, por_mes, resumen_global)

    detalle = [{
        "semana":        r["semana"],
        "mes":           r.get("mes"),
        "internista_norm":   _norm_internista(r["medico"]),
        "transcriptor_norm": _norm_transcriptor(r["transcriptor"]),
        "realizacion":   r["realizacion"].isoformat() if r["realizacion"] else None,
        "semaforo":      r["semaforo"].isoformat() if r["semaforo"] else None,
        "terminado":     r["terminado"].isoformat() if r["terminado"] else None,
        "entregado":     r["entregado"].isoformat() if r["entregado"] else None,
        "medico":        r["medico"],
        "transcriptor":  r["transcriptor"],
        "estado":        r["estado"],
        "entrega_digital": r["entrega_digital"],
        "d_realiz_terminado":    r["d_realiz_terminado"],
        "d_terminado_entregado": r["d_terminado_entregado"],
        "d_realiz_entregado":    r["d_realiz_entregado"],
        "d_realiz_semaforo":     r["d_realiz_semaforo"],
        "d_entregado_vs_real":   r["d_entregado_vs_real"],
    } for r in registros]

    semanas = sorted({str(r["semana"]) for r in registros if r["semana"]})

    return {
        "procesos":       procesos,
        "transcriptores": transcriptores,
        "distribucion":   distribucion,
        "total":          len(registros),
        "entregados":     entregados,
        "en_proceso":     len(registros) - entregados,
        "lim_verde":      CHK_LIM_VERDE,
        "lim_amarillo":   CHK_LIM_AMARILLO,
        "semanas":        semanas,
        "detalle":        detalle,
        # ── Métricas solicitadas (acumulado + desglose) ──────────────────────
        "resumen":        resumen_global,
        "por_semana":     por_semana,
        "por_mes":        por_mes,
        "lista_semanas":  [str(s) for s in semanas_ord],
        "lista_meses":    meses_ord,
        "obj_entrega_real": OBJ_ENTREGA_REAL,
        "obj_realiz_term":  OBJ_REALIZ_TERM,
        "lim_realiz_term":  LIM_REALIZ_TERM,
        "patrones":       patrones,
    }


def _analisis_web(rows):
    """Calcula todos los análisis detallados para el dashboard web."""
    from collections import defaultdict, Counter

    def avg(lst): return sum(lst) / len(lst) if lst else None

    DEPTOS_MODAL = set(DEPTO_A_MOD.keys())

    # ── Incumplimiento por departamento ──────────────────────────────────────
    depto_data = defaultdict(lambda: {"total": 0, "incump": 0, "estudios_incump": Counter()})
    for r in rows:
        d = r["depto"]
        depto_data[d]["total"] += 1
        if es_incumplimiento(r["af"]):
            depto_data[d]["incump"] += 1
            depto_data[d]["estudios_incump"][r["estudio"]] += 1

    incump_depto = []
    for d, v in sorted(depto_data.items(), key=lambda x: -x[1]["incump"]):
        top_est = [{"estudio": e, "n": n}
                   for e, n in v["estudios_incump"].most_common(5)]
        incump_depto.append({
            "depto":   d,
            "total":   v["total"],
            "incump":  v["incump"],
            "pct_cumpl": round((v["total"] - v["incump"]) / v["total"] * 100, 1)
                         if v["total"] else None,
            "top_estudios": top_est,
        })

    # ── Top estudios con más incumplimientos (global) ────────────────────────
    est_incump = Counter(r["estudio"] for r in rows if es_incumplimiento(r["af"]))
    top_estudios_incump = [{"estudio": e, "n": n}
                           for e, n in est_incump.most_common(10)]

    # ── Mapa de calor: depto × sucursal → %cumplimiento ─────────────────────
    SUCURSALES_ORD = list(SUCURSALES.keys())
    deptos_modal   = sorted(DEPTOS_MODAL)
    calor = {}
    for d in deptos_modal:
        calor[d] = {}
        for cod, nom in SUCURSALES.items():
            sub = [r for r in rows if r["depto"] == d and r["sucursal"] == nom]
            if not sub:
                calor[d][cod] = None
            else:
                inc = sum(1 for r in sub if es_incumplimiento(r["af"]))
                calor[d][cod] = round((len(sub) - inc) / len(sub) * 100, 1)

    # ── Tiempos de proceso por modalidad (GC y GF) ───────────────────────────
    tiempos_modal = {}
    for mod, cfg in MODALIDADES.items():
        depto  = cfg["depto"]
        lim_ag = cfg["lim_ag"]
        sub    = [r for r in rows if r["depto"] == depto
                  and r["ag"] is not None and r["ag"] <= lim_ag]
        gc_vals = [r["gc"] for r in sub if r["gc"] is not None and 0 < r["gc"] <= lim_ag * 3]
        gf_vals = [r["gf"] for r in sub if r["gf"] is not None and 0 < r["gf"] <= lim_ag * 3]
        tiempos_modal[mod] = {
            "gc_avg": round(avg(gc_vals), 2) if avg(gc_vals) else None,
            "gf_avg": round(avg(gf_vals), 2) if avg(gf_vals) else None,
            "n":      len(sub),
        }

    # ── Usuarios recibe entregas ──────────────────────────────────────────────
    usr_data = defaultdict(lambda: {"total": 0, "incump": 0, "deptos": Counter()})
    for r in rows:
        u = r.get("usuario_recibe", "").strip()
        if not u:
            continue
        usr_data[u]["total"] += 1
        usr_data[u]["deptos"][r["depto"]] += 1
        if es_incumplimiento(r["af"]):
            usr_data[u]["incump"] += 1

    usuarios_recibe = sorted(
        [{"usuario": u,
          "total":   v["total"],
          "incump":  v["incump"],
          "pct_cumpl": round((v["total"] - v["incump"]) / v["total"] * 100, 1)
                       if v["total"] else None,
          "top_deptos": [{"depto": d, "n": n}
                         for d, n in v["deptos"].most_common(3)]}
         for u, v in usr_data.items()],
        key=lambda x: -x["total"]
    )

    # ── Distribución A-F de incumplimientos (histograma en intervalos de 1h) ─
    incump_rows = [r for r in rows if es_incumplimiento(r["af"])]
    buckets = defaultdict(int)
    for r in incump_rows:
        b = int(abs(r["af"]))
        buckets[b] += 1
    dist_af = [{"hora": h, "n": buckets.get(h, 0)} for h in range(24)]

    # ── Mapa de calor horario: hora (7-21) × departamento ────────────────────
    HORAS = list(range(7, 22))
    deptos_incump = sorted({r["depto"] for r in incump_rows})
    calor_hora = {}
    for d in deptos_incump:
        calor_hora[d] = {}
        for h in HORAS:
            calor_hora[d][h] = sum(
                1 for r in incump_rows
                if r["depto"] == d and r.get("hora_alta") == h
            )
    # Total por hora (todos los deptos)
    total_hora = {h: sum(1 for r in incump_rows if r.get("hora_alta") == h)
                  for h in HORAS}

    return {
        "incump_depto":        incump_depto,
        "top_estudios_incump": top_estudios_incump,
        "calor_deptos":        deptos_modal,
        "calor_sucursales":    SUCURSALES_ORD,
        "calor":               calor,
        "tiempos_modal":       tiempos_modal,
        "usuarios_recibe":     usuarios_recibe,
        "dist_af":             dist_af,
        "calor_hora":          calor_hora,
        "calor_hora_total":    total_hora,
        "calor_hora_deptos":   deptos_incump,
        "calor_horas":         HORAS,
    }


def exportar_web(tabla_acum, tabla_por_semana, todas_las_filas,
                 semanas_presentes, fecha_min, fecha_max):
    """Genera docs/data.js con los datos para el dashboard web (GitHub Pages)."""
    def clean(obj):
        if isinstance(obj, dict):
            return {k: clean(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [clean(v) for v in obj]
        if isinstance(obj, datetime.date):
            return obj.isoformat()
        if isinstance(obj, float) and obj != obj:
            return None
        return obj

    # Rango de fechas por semana
    rangos = {}
    for sem in semanas_presentes:
        fechas = [r["fecha"] for r in todas_las_filas
                  if r.get("semana") == sem and r["fecha"]]
        if fechas:
            rangos[str(sem)] = {"inicio": min(fechas).isoformat(),
                                "fin": max(fechas).isoformat()}

    # Análisis detallados acumulados y por semana
    analisis_acum = _analisis_web(todas_las_filas)
    analisis_sem  = {}
    for sem in semanas_presentes:
        sub = [r for r in todas_las_filas if r.get("semana") == sem]
        analisis_sem[str(sem)] = _analisis_web(sub)

    data = {
        "generado":        datetime.datetime.now().isoformat(),
        "rango":           {"inicio": fecha_min.isoformat(), "fin": fecha_max.isoformat()},
        "semanas":         sorted(str(s) for s in semanas_presentes),
        "rangos":          rangos,
        "acumulado":       tabla_acum,
        "por_semana":      tabla_por_semana,
        "sucursales":      SUCURSALES,
        "total_registros": len(todas_las_filas),
        "analisis":        analisis_acum,
        "analisis_sem":    analisis_sem,
        "interpretaciones": leer_todas_interpretaciones(),
        "chequeos":        _analisis_chequeos(leer_todos_chequeos()),
    }

    DOCS_DIR = Path(__file__).parent / "docs"
    DOCS_DIR.mkdir(exist_ok=True)
    ruta = DOCS_DIR / "data.js"
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("// Generado automáticamente por procesar.py — no editar a mano\n")
        f.write("window.DASHBOARD_DATA = ")
        json.dump(clean(data), f, ensure_ascii=False, indent=2, default=str)
        f.write(";\n")
    return ruta


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Dashboard de Entrega de Resultados")
    parser.add_argument("--semana",   type=int, nargs="+",
                        help="Número(s) de semana a procesar (default: todas)")
    parser.add_argument("--archivo",  nargs="+",
                        help="Archivos XLS específicos a procesar")
    parser.add_argument("--salida",   default=None,
                        help="Nombre del archivo de salida (sin extensión)")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)

    # Determinar archivos a leer
    if args.archivo:
        archivos = [Path(a) for a in args.archivo]
    else:
        archivos = sorted(DATOS_DIR.glob("*.xls")) + sorted(DATOS_DIR.glob("*.xlsx"))

    if not archivos:
        print("No se encontraron archivos XLS en datos/. "
              "Coloca los archivos ahí o usa --archivo.")
        sys.exit(1)

    # Filtrar por semana si se especificó
    if args.semana:
        archivos = [a for a in archivos
                    if semana_de_archivo(a) in args.semana or semana_de_archivo(a) is None]

    print(f"Procesando {len(archivos)} archivo(s):")
    todas_las_filas = []
    semanas_presentes = set()
    for a in archivos:
        print(f"  • {a.name}", end="  ")
        filas = leer_xls(a)
        sem   = semana_de_archivo(a)
        if sem:
            semanas_presentes.add(sem)
        print(f"({len(filas)} registros)")
        todas_las_filas.extend(filas)

    if not todas_las_filas:
        print("No se obtuvieron registros válidos.")
        sys.exit(1)

    # Fechas para nombre de archivo
    fechas_validas = [r["fecha"] for r in todas_las_filas if r["fecha"]]
    fecha_max  = max(fechas_validas) if fechas_validas else datetime.date.today()
    fecha_min  = min(fechas_validas) if fechas_validas else fecha_max

    # Título del reporte
    if semanas_presentes:
        sems_str = ", ".join(f"Sem {s}" for s in sorted(semanas_presentes))
        titulo   = f"Dashboard de Entrega de Resultados  —  {sems_str}  ({fecha_min} al {fecha_max})"
    else:
        titulo   = f"Dashboard de Entrega de Resultados  —  {fecha_min} al {fecha_max}"

    # Tabla acumulada
    tabla_acum = generar_tabla(todas_las_filas, fecha_max)

    # Tabla por semana
    tabla_por_semana = {}
    for sem in sorted(semanas_presentes):
        rows_sem = [r for r in todas_las_filas if r.get("semana") == sem]
        tabla_por_semana[str(sem)] = generar_tabla(rows_sem)

    # Nombre de salida
    sufijo = args.salida or f"Dashboard_{fecha_max.strftime('%d_%m_%Y')}"
    if args.semana and len(args.semana) == 1:
        sufijo = f"Dashboard_Semana{args.semana[0]}_{fecha_max.strftime('%d_%m_%Y')}"

    ruta_xlsx = OUTPUT_DIR / f"{sufijo}.xlsx"
    ruta_json = OUTPUT_DIR / f"{sufijo}.json"

    # Generar Excel
    wb = openpyxl.Workbook()

    # Hoja 1: Acumulado
    ws_acum = wb.active
    ws_acum.title = "Acumulado"
    escribir_hoja_dashboard(ws_acum, titulo, tabla_acum)

    # Hoja 2: Por semana (si hay más de una o si se quiere el desglose)
    if semanas_presentes:
        ws_sem = wb.create_sheet("Por Semana")
        escribir_hoja_semanas(ws_sem, todas_las_filas, list(semanas_presentes))

    # Hoja 3: Detalle de registros
    ws_det = wb.create_sheet("Detalle Registros")
    escribir_hoja_detalle(ws_det, todas_las_filas)

    # Hoja 4: Días por proceso (chequeos de transcripción / entrega)
    chequeos = leer_todos_chequeos()
    if chequeos:
        ws_chk = wb.create_sheet("Días por Proceso")
        escribir_hoja_chequeos(ws_chk, chequeos)
        print(f"  • Chequeos: {len(chequeos)} registros → hoja 'Días por Proceso'")

    wb.save(ruta_xlsx)
    print(f"\n✓ Excel generado: {ruta_xlsx}")

    # Generar JSON
    exportar_json(tabla_acum, tabla_por_semana, todas_las_filas, ruta_json)
    print(f"✓ JSON  generado: {ruta_json}")

    # Generar datos para el dashboard web (GitHub Pages)
    ruta_web = exportar_web(tabla_acum, tabla_por_semana, todas_las_filas,
                            semanas_presentes, fecha_min, fecha_max)
    print(f"✓ Web   generado: {ruta_web}")

    # Resumen rápido en consola
    general = tabla_acum[0]
    print(f"\n── Resumen Acumulado ──────────────────────────")
    print(f"  Registros totales :  {general['modalidades_total']}")
    print(f"  Estudios modal.   :  {general['estudios']}")
    print(f"  Incumplimientos   :  {general['incumplimiento']}")
    ind = general["indicador"]
    print(f"  Indicador global  :  {ind:.2f}%" if ind else "  Indicador global  :  -")
    print(f"────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()
